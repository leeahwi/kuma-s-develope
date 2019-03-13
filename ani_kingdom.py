import discord
import asyncio
import openpyxl
import random
import os


client = discord.Client()


@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("ㅡㅡㅡㅡㅡㅡ")
    await client.change_presence(game=discord.Game(name='테스트 중', type=1))

@client.event
async def on_message(message):
    if message.content.startswith("!안녕"):
        await client.send_message(message.channel, "안녕하세요")
    if message.content.startswith("!보스"):
        await client.send_message(message.channel, "조용히 하세요")
    if message.content.startswith("!희람"):
        await client.send_message(message.channel, ":pig:")

    if message.content.startswith("!팜팁1"):
        await client.send_file(message.channel, 'image/팜팁1.gif')
    if message.content.startswith("!에반데"):
        await client.send_file(message.channel, 'image/에반데1.jpg')
    if message.content.startswith("!천팔백육십삼진에바"):
        await client.send_file(message.channel, 'image/에반데2.jpg')



    if message.content.startswith("!메모") and not message.content.startswith("!메모 지우기") and not message.content.startswith("!메모 청소하기"):
        file = openpyxl.load_workbook("메모장.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await  client.send_message(message.channel, "done")
                break

            elif sheet["A" + str(50)].value != "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await  client.send_message(message.channel, "done")
                break
        file.save("메모장.xlsx")

    if message.content.startswith("!설명"):
        file = openpyxl.load_workbook("메모장.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                if sheet["C" + str(i)].value == "-":
                    await  client.send_message(message.channel, sheet["B" + str(i)].value)
                    break
                elif sheet["D" + str(i)].value == "-":
                    await  client.send_message(message.channel, sheet["B" + str(i)].value + " " + sheet["C" + str(i)].value)
                    break
                else:
                    await  client.send_message(message.channel,
                            sheet["B" + str(i)].value + " " + sheet["C" + str(i)].value
                                                      + " " + sheet["D" + str(i)].value)

    if message.content.startswith("!메모 지우기"):
        file = openpyxl.load_workbook("메모장.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                sheet["C" + str(i)].value = "-"
                sheet["D" + str(i)].value = "-"
                await client.send_message(message.channel, "해당 메모 정리 끝")
                file.save("메모장.xlsx")
                break

    if message.content.startswith("!메모 청소하기"):
        file = openpyxl.load_workbook("메모장.xlsx")
        sheet = file.active
        for i in range(1, 51):
            sheet["A" + str(i)].value = "-"
            sheet["B" + str(i)].value = "-"
            sheet["C" + str(i)].value = "-"
            sheet["D" + str(i)].value = "-"
        await client.send_message(message.channel, "모든 메모 정리 끝")
        file.save("메모장.xlsx")






access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
